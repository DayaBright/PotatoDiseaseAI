"""
PASO 2 — Auditoría Completa y Análisis Cruzado de Datasets

Funciones:
  1. Escaneo de los 7 datasets sin duplicar extensiones (.png vs .PNG)
  2. Detección de imágenes corruptas
  3. Blur normalizado por resolución (para ver si una imagen es borrosa)
  4. Detección de duplicados cruzados entre todos los datasets
  5. Detección automática de subconjuntos (DS_A ⊂ DS_B)
  6. Fungi y virus marcados como __REVISAR__ (pueden tener errores de etiquetado)
  7. Grillas de 25 imágenes por clase por dataset (para revisión visual)
  8. Excel de duplicados para revisión manual
  9. Reporte HTML autocontenido único
  10. CSV de inventario completo

Dependencias:
    pip install Pillow numpy matplotlib imagehash openpyxl
"""

import base64
import csv
import io
import random
import hashlib
import logging
import sys
from collections import defaultdict
from datetime import datetime
from itertools import combinations
from pathlib import Path

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from PIL import Image, ImageFilter, UnidentifiedImageError

# ─── DEPENDENCIAS OPCIONALES ────────────────────────────────────────────────
try:
    import imagehash
    USE_PHASH = True
except ImportError:
    USE_PHASH = False
    print("[AVISO] imagehash no instalado — se usará MD5.")
    print("        pip install imagehash")

try:
    import openpyxl
    USE_XLSX = True
except ImportError:
    USE_XLSX = False
    print("[AVISO] openpyxl no instalado — no se generará Excel.")
    print("        pip install openpyxl")

# ════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN
# ════════════════════════════════════════════════════════════════════════════

DATASETS = {
    "dataset1_warcoder":        r"C:\Users\jaevi\Downloads\Datasets\dataset curado\dataset1_warcoder",
    "dataset2_rgfhz":           r"C:\Users\jaevi\Downloads\Datasets\dataset curado\dataset2_rgfhz",
    "dataset3_nirmalsankalana": r"C:\Users\jaevi\Downloads\Datasets\dataset curado\dataset3_nirmalsankalana",
    "dataset4_NoDescription":   r"C:\Users\jaevi\Downloads\Datasets\dataset curado\dataset4_NoDescription",
    "dataset5_shahadhossin":    r"C:\Users\jaevi\Downloads\Datasets\dataset curado\dataset5_shahadhossin",
    "dataset6_markkostantine":  r"C:\Users\jaevi\Downloads\Datasets\dataset curado\dataset6_markkostantine",
    "dataset7_nirmalsankalana": r"C:\Users\jaevi\Downloads\Datasets\dataset curado\dataset7_nirmalsankalana",
}

OUTPUT_DIR     = Path("auditoria_output")
INVENTORY_CSV  = OUTPUT_DIR / "inventario_imagenes.csv"
REPORT_HTML    = OUTPUT_DIR / "reporte_auditoria.html"
DUPLICATES_XLS = OUTPUT_DIR / "duplicados_revision.xlsx"

VALID_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".tiff", ".tif"}

# Muestras por clase por dataset en las grillas visuales
N_MUESTRAS = 25

# Umbral para declarar subconjunto:
# si DS_A tiene >= SUBSET_THRESHOLD% de sus hashes en DS_B → DS_A ⊂ DS_B
SUBSET_THRESHOLD = 80.0

RANDOM_SEED = 42
random.seed(RANDOM_SEED)
np.random.seed(RANDOM_SEED)

# ─── MAPEO DE ETIQUETAS ──────────────────────────────────────────────────────
# fungi y virus → __REVISAR__ (revisión manual antes de consolidar)
# Claves en minúsculas con guión bajo (normalize_label hace esa conversión)
LABEL_MAP = {
    # late_blight
    "late_blight":              "late_blight",
    "potato___late_blight":     "late_blight",
    "phytopthora":              "late_blight",   # typo original DS1/DS6
    "pyhtopora":                "late_blight",   # typo original DS6
    "potato___phytophthora":    "late_blight",

    # early_blight
    "early_blight":             "early_blight",
    "potato___early_blight":    "early_blight",

    # leafroll_virus
    "plrv_leaf":                "leafroll_virus",
    "potato___leafroll_virus":  "leafroll_virus",

    # mosaic_virus
    "mosaic_leaf":              "mosaic_virus",
    "potato___mosaic_virus":    "mosaic_virus",

    # bacterial_wilt
    "bacteria":                 "bacterial_wilt",
    "potato___bacterial_wilt":  "bacterial_wilt",

    # nematode
    "nematode":                 "nematode",
    "potato_cyst_nematode":     "nematode",
    "potato___nematode":        "nematode",

    # pest
    "pest":                     "pest",
    "plant_pests":              "pest",
    "potato___pests":           "pest",

    # healthy
    "healthy":                  "healthy",
    "healthy_leaf_images":      "healthy",
    "potato___healthy":         "healthy",

    # __REVISAR__ — fungi y virus genéricos para clasificación manual
    "fungi":                    "__REVISAR__",
    "fungal_diseases":          "__REVISAR__",
    "virus":                    "__REVISAR__",
    "potato_virus":             "__REVISAR__",
}

TARGET_CLASSES = [
    "late_blight", "early_blight", "leafroll_virus", "mosaic_virus",
    "bacterial_wilt", "nematode", "pest", "healthy",
]

# ─── LOGGING ────────────────────────────────────────────────────────────────
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(OUTPUT_DIR / "auditoria.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# ════════════════════════════════════════════════════════════════════════════
# FUNCIONES AUXILIARES
# ════════════════════════════════════════════════════════════════════════════

def normalize_label(folder_name: str) -> str:
    """Normaliza nombre de carpeta a clave del LABEL_MAP."""
    key = folder_name.strip().lower().replace(" ", "_")
    if key in LABEL_MAP:
        return LABEL_MAP[key]
    # Búsqueda parcial como fallback
    for pattern, target in LABEL_MAP.items():
        if pattern in key or key in pattern:
            return target
    return "__DESCONOCIDO__"


def is_valid_image(path: Path) -> tuple:
    """Verifica integridad de la imagen. Retorna (bool, mensaje)."""
    try:
        with Image.open(path) as img:
            img.verify()
        return True, ""
    except Exception as e:
        return False, str(e)


def compute_hash(path: Path) -> str:
    """pHash perceptual si imagehash disponible, MD5 si no."""
    if USE_PHASH:
        try:
            with Image.open(path) as img:
                return str(imagehash.phash(img))
        except Exception:
            return "ERROR"
    else:
        md5 = hashlib.md5()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""):
                md5.update(chunk)
        return md5.hexdigest()


def blur_score_normalized(path: Path) -> float:
    """
    Varianza del Laplaciano normalizada por resolución.
    Divide por (width * height) para que sea comparable entre
    imágenes de 224px y de 4000px.
    Retorna score normalizado — solo informativo, no excluye.
    """
    try:
        with Image.open(path).convert("L") as img:
            w, h = img.size
            edges = img.filter(ImageFilter.FIND_EDGES)
            arr = np.array(edges, dtype=np.float32)
            raw_var = float(np.var(arr))
            # Normalizar por área para comparar entre resoluciones
            normalized = raw_var / (w * h) * 1e6
            return round(normalized, 4)
    except Exception:
        return 0.0


def get_resolution(path: Path):
    try:
        with Image.open(path) as img:
            return img.size  # (width, height)
    except Exception:
        return (0, 0)


def img_to_base64(path: str, size: tuple = (100, 100)) -> str:
    """Carga imagen, redimensiona y retorna base64 para incrustar en HTML."""
    try:
        with Image.open(path) as img:
            img = img.convert("RGB")
            img.thumbnail(size, Image.LANCZOS)
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=70)
            return base64.b64encode(buf.getvalue()).decode("utf-8")
    except Exception:
        return ""


# ════════════════════════════════════════════════════════════════════════════
# PASO 1 — ESCANEO DE DATASETS
# ════════════════════════════════════════════════════════════════════════════

def scan_all_datasets() -> list:
    """
    Escanea los 7 datasets y retorna lista de registros.
    Fix Windows: usa resolve() + set para evitar contar
    la misma imagen dos veces por diferencia de mayúsculas.
    """
    all_records = []
    corrupt_list = []

    for ds_name, ds_path in DATASETS.items():
        root = Path(ds_path)
        if not root.exists():
            log.error(f"[{ds_name}] Ruta no existe: {root}")
            continue

        log.info(f"\n{'='*60}")
        log.info(f"Escaneando: {ds_name}")

        # Deduplicar rutas en Windows (case-insensitive filesystem)
        seen = set()
        image_paths = []
        for candidate in root.rglob("*"):
            if candidate.suffix.lower() in VALID_EXTENSIONS:
                resolved = str(candidate.resolve()).lower()
                if resolved not in seen:
                    seen.add(resolved)
                    image_paths.append(candidate)

        log.info(f"  Archivos encontrados: {len(image_paths)}")
        unknown_folders = set()

        for img_path in image_paths:
            parent = img_path.parent.name
            mapped = normalize_label(parent)

            if mapped == "__DESCONOCIDO__":
                unknown_folders.add(parent)

            # Integridad
            valid, err = is_valid_image(img_path)
            if not valid:
                corrupt_list.append({
                    "dataset": ds_name, "path": str(img_path), "error": err
                })
                log.warning(f"  CORRUPTA: {img_path.name} — {err}")
                continue

            w, h = get_resolution(img_path)
            blur  = blur_score_normalized(img_path)
            hash_ = compute_hash(img_path)

            all_records.append({
                "dataset":         ds_name,
                "path":            str(img_path),
                "filename":        img_path.name,
                "original_folder": parent,
                "mapped_class":    mapped,
                "width":           w,
                "height":          h,
                "blur_norm":       blur,
                "hash":            hash_,
            })

        if unknown_folders:
            log.warning(f"  Carpetas no mapeadas: {unknown_folders}")

    log.info(f"\nTotal registros escaneados: {len(all_records)}")
    log.info(f"Total imágenes corruptas  : {len(corrupt_list)}")
    return all_records, corrupt_list


# ════════════════════════════════════════════════════════════════════════════
# PASO 2 — DUPLICADOS CRUZADOS ENTRE TODOS LOS DATASETS
# ════════════════════════════════════════════════════════════════════════════

def detect_cross_duplicates(records: list) -> dict:
    """
    Agrupa imágenes por hash considerando TODOS los datasets juntos.
    Clasifica cada grupo como:
      - cross_dataset: misma imagen en ≥2 datasets distintos
      - cross_class:   misma imagen en ≥2 clases distintas (error de etiquetado)
      - intra:         copia dentro del mismo dataset y clase
    """
    hash_groups = defaultdict(list)
    for r in records:
        h = r.get("hash", "")
        if h and h != "ERROR":
            hash_groups[h].append(r)

    duplicates = {}
    for h, entries in hash_groups.items():
        if len(entries) < 2:
            continue
        datasets = {e["dataset"]      for e in entries}
        classes  = {e["mapped_class"] for e in entries}
        duplicates[h] = {
            "entries":       entries,
            "cross_dataset": len(datasets) > 1,
            "cross_class":   len(classes)  > 1,
            "datasets":      sorted(datasets),
            "classes":       sorted(classes),
        }

    cross_ds  = sum(1 for d in duplicates.values() if d["cross_dataset"])
    cross_cls = sum(1 for d in duplicates.values() if d["cross_class"] and not d["cross_dataset"])
    intra     = len(duplicates) - cross_ds - cross_cls

    log.info(f"\n[Duplicados]")
    log.info(f"  Entre datasets distintos : {cross_ds}")
    log.info(f"  Entre clases distintas   : {cross_cls}  ← revisar etiquetado")
    log.info(f"  Intra-clase / intra-DS   : {intra}")
    log.info(f"  Total grupos             : {len(duplicates)}")

    return duplicates


# ════════════════════════════════════════════════════════════════════════════
# PASO 3 — DETECCIÓN DE SUBCONJUNTOS
# ════════════════════════════════════════════════════════════════════════════

def detect_subsets(records: list) -> list:
    """
    Para cada par de datasets calcula el porcentaje de hashes compartidos.
    Si DS_A tiene >= SUBSET_THRESHOLD% de sus hashes en DS_B → DS_A ⊂ DS_B.
    Detecta cualquier relación, no solo las conocidas (DS1/DS6, DS2/DS7).
    """
    ds_hashes = defaultdict(set)
    for r in records:
        h = r.get("hash", "")
        if h and h != "ERROR":
            ds_hashes[r["dataset"]].add(h)

    relations = []
    for ds_a, ds_b in combinations(sorted(ds_hashes.keys()), 2):
        ha = ds_hashes[ds_a]
        hb = ds_hashes[ds_b]
        if not ha or not hb:
            continue

        shared  = len(ha & hb)
        pct_a   = shared / len(ha) * 100  # % de A contenido en B
        pct_b   = shared / len(hb) * 100  # % de B contenido en A

        if pct_a >= SUBSET_THRESHOLD or pct_b >= SUBSET_THRESHOLD:
            if pct_a >= pct_b:
                subset, superset, pct = ds_a, ds_b, pct_a
            else:
                subset, superset, pct = ds_b, ds_a, pct_b

            relations.append({
                "subset":    subset,
                "superset":  superset,
                "pct":       round(pct, 1),
                "shared":    shared,
                "total_sub": len(ds_hashes[subset]),
                "total_sup": len(ds_hashes[superset]),
            })
            log.info(f"\n[SUBCONJUNTO] {subset} ⊂ {superset}")
            log.info(f"  {pct:.1f}% de {subset} encontrado en {superset}")
            log.info(f"  Imágenes compartidas: {shared}")
            log.info(f"  → EXCLUIR {subset}")

    relations.sort(key=lambda x: x["pct"], reverse=True)

    if not relations:
        log.info("\n[OK] No se detectaron relaciones de subconjunto.")

    return relations


# ════════════════════════════════════════════════════════════════════════════
# PASO 4 — GRILLAS VISUALES (25 muestras por clase por dataset)
# ════════════════════════════════════════════════════════════════════════════

def generate_grids(records: list) -> dict:
    """
    Para cada (dataset, clase) genera lista de imágenes en base64.
    Incluye clases TARGET + __REVISAR__ + __DESCONOCIDO__.
    """
    groups = defaultdict(list)
    for r in records:
        groups[(r["dataset"], r["mapped_class"])].append(r["path"])

    grids = {}
    total = len(groups)
    for i, ((ds, cls), paths) in enumerate(sorted(groups.items())):
        sample = random.sample(paths, min(N_MUESTRAS, len(paths)))
        b64_list = [img_to_base64(p) for p in sample]
        b64_list = [b for b in b64_list if b]  # filtrar errores
        grids[(ds, cls)] = b64_list
        log.info(f"  Grilla {i+1}/{total}: {ds} / {cls} — {len(b64_list)} imgs")

    return grids


# ════════════════════════════════════════════════════════════════════════════
# PASO 5 — CSV DE INVENTARIO
# ════════════════════════════════════════════════════════════════════════════

def save_csv(records: list):
    fields = [
        "dataset", "path", "filename", "original_folder",
        "mapped_class", "width", "height", "blur_norm", "hash"
    ]
    with open(INVENTORY_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for r in records:
            writer.writerow({k: r[k] for k in fields})
    log.info(f"[OK] CSV guardado: {INVENTORY_CSV} ({len(records)} filas)")


# ════════════════════════════════════════════════════════════════════════════
# PASO 6 — EXCEL DE DUPLICADOS
# ════════════════════════════════════════════════════════════════════════════

def save_duplicates_excel(duplicates: dict):
    if not USE_XLSX:
        log.warning("[OMITIDO] openpyxl no disponible.")
        return

    wb = openpyxl.Workbook()

    # Hoja 1: duplicados entre datasets
    ws1 = wb.active
    ws1.title = "Entre_datasets"
    ws1.append([
        "Hash", "Tipo", "Dataset_A", "Clase_A", "Ruta_A",
        "Dataset_B", "Clase_B", "Ruta_B", "Accion_sugerida"
    ])

    # Hoja 2: duplicados entre clases (posible error de etiquetado)
    ws2 = wb.create_sheet("Error_etiquetado")
    ws2.append([
        "Hash", "Clase_1", "Dataset_1", "Ruta_1",
        "Clase_2", "Dataset_2", "Ruta_2", "Accion_sugerida"
    ])

    # Hoja 3: duplicados intra-clase
    ws3 = wb.create_sheet("Intra_clase")
    ws3.append([
        "Hash", "Dataset", "Clase", "Ruta_1", "Ruta_2", "Accion_sugerida"
    ])

    for h, info in duplicates.items():
        entries = info["entries"]
        pairs = [
            (entries[i], entries[j])
            for i in range(len(entries))
            for j in range(i + 1, len(entries))
        ]

        for e1, e2 in pairs:
            cross_ds  = e1["dataset"]      != e2["dataset"]
            cross_cls = e1["mapped_class"] != e2["mapped_class"]

            if cross_ds and cross_cls:
                ws2.append([
                    h,
                    e1["mapped_class"], e1["dataset"], e1["path"],
                    e2["mapped_class"], e2["dataset"], e2["path"],
                    "REVISAR — misma imagen en clases distintas, posible error de etiquetado"
                ])
            elif cross_ds:
                ws1.append([
                    h, "CROSS-DS",
                    e1["dataset"], e1["mapped_class"], e1["path"],
                    e2["dataset"], e2["mapped_class"], e2["path"],
                    "Conservar en el superconjunto — se elimina al excluir el subconjunto"
                ])
            else:
                ws3.append([
                    h,
                    e1["dataset"], e1["mapped_class"],
                    e1["path"], e2["path"],
                    "Conservar uno, eliminar el duplicado"
                ])

    # Ajustar anchos
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 70)

    wb.save(DUPLICATES_XLS)
    log.info(f"[OK] Excel guardado: {DUPLICATES_XLS}")


# ════════════════════════════════════════════════════════════════════════════
# PASO 7 — REPORTE HTML AUTOCONTENIDO
# ════════════════════════════════════════════════════════════════════════════

def save_html_report(records, corrupt_list, duplicates, subsets, grids):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── Conteos por clase ────────────────────────────────────────────────────
    class_totals = defaultdict(int)
    for r in records:
        class_totals[r["mapped_class"]] += 1

    total_entrenables = sum(class_totals.get(c, 0) for c in TARGET_CLASSES)

    # ── Sección resumen de clases ────────────────────────────────────────────
    summary_rows = ""
    for cls in TARGET_CLASSES:
        n = class_totals.get(cls, 0)
        pct = round(n / total_entrenables * 100, 1) if total_entrenables else 0
        if n >= 1000:
            color, estado = "#d4edda", "✔ OK"
        elif n >= 400:
            color, estado = "#fff3cd", "⚠ Aceptable"
        else:
            color, estado = "#f8d7da", "✘ Crítico"
        summary_rows += f"""
        <tr style="background:{color}">
            <td><b>{cls}</b></td>
            <td>{n:,}</td>
            <td>{pct}%</td>
            <td>{estado}</td>
        </tr>"""

    revisar_n = class_totals.get("__REVISAR__", 0)
    descon_n  = class_totals.get("__DESCONOCIDO__", 0)
    summary_rows += f"""
        <tr style="background:#fff3cd">
            <td><b>__REVISAR__</b> (fungi / virus)</td>
            <td>{revisar_n:,}</td><td>—</td>
            <td>⚠ Revisión manual requerida</td>
        </tr>
        <tr style="background:#f8d7da">
            <td><b>__DESCONOCIDO__</b></td>
            <td>{descon_n:,}</td><td>—</td>
            <td>✘ Añadir al LABEL_MAP</td>
        </tr>"""

    # ── Sección subconjuntos ─────────────────────────────────────────────────
    if subsets:
        sub_rows = ""
        for s in subsets:
            cert_color = "#d4edda" if s["pct"] >= 95 else "#fff3cd"
            cert_label = "Alta certeza" if s["pct"] >= 95 else "Revisar"
            sub_rows += f"""
            <tr style="background:{cert_color}">
                <td><b>{s['subset']}</b></td>
                <td>{s['superset']}</td>
                <td>{s['pct']}%</td>
                <td>{s['shared']:,}</td>
                <td>{s['total_sub']:,}</td>
                <td>{s['total_sup']:,}</td>
                <td>{cert_label}</td>
                <td style="color:#c00"><b>EXCLUIR {s['subset']}</b></td>
            </tr>"""
        subset_section = f"""
        <h2>Subconjuntos detectados</h2>
        <div class="warn">
            Umbral: ≥{SUBSET_THRESHOLD}% de hashes compartidos.
            Alta certeza = ≥95%. Estos datasets deben excluirse antes de consolidar.
        </div>
        <table>
            <tr>
                <th>Subconjunto (excluir)</th><th>Superconjunto (conservar)</th>
                <th>% compartido</th><th>Imgs compartidas</th>
                <th>Total sub</th><th>Total super</th>
                <th>Certeza</th><th>Decisión</th>
            </tr>{sub_rows}
        </table>"""
    else:
        subset_section = """
        <h2>Subconjuntos</h2>
        <div class="info">No se detectaron relaciones de subconjunto.</div>"""

    # ── Sección duplicados ───────────────────────────────────────────────────
    cross_ds  = sum(1 for d in duplicates.values() if d["cross_dataset"] and not d["cross_class"])
    cross_cls = sum(1 for d in duplicates.values() if d["cross_class"])
    intra     = sum(1 for d in duplicates.values() if not d["cross_dataset"] and not d["cross_class"])

    dup_section = f"""
    <h2>Duplicados detectados</h2>
    <div class="info">
        Método: {'pHash perceptual' if USE_PHASH else 'MD5 exacto'}.
        Ver <code>duplicados_revision.xlsx</code> para revisión detallada.
    </div>
    <table>
        <tr><th>Tipo</th><th>Grupos</th><th>Descripción</th><th>Acción</th></tr>
        <tr style="background:#f8d7da">
            <td>Entre datasets distintos</td><td><b>{cross_ds}</b></td>
            <td>Misma imagen en ≥2 datasets</td>
            <td>Se resuelve al excluir los subconjuntos</td>
        </tr>
        <tr style="background:#f8d7da">
            <td>Entre clases distintas</td><td><b>{cross_cls}</b></td>
            <td>Misma imagen en ≥2 clases — posible error de etiquetado</td>
            <td>Revisar hoja "Error_etiquetado" del Excel</td>
        </tr>
        <tr style="background:#fff3cd">
            <td>Intra-clase / intra-dataset</td><td><b>{intra}</b></td>
            <td>Copias dentro del mismo dataset y clase</td>
            <td>Conservar una, eliminar el resto</td>
        </tr>
    </table>"""

    # ── Sección blur (solo informativa) ─────────────────────────────────────
    blur_section = """
    <h2>Blur normalizado (solo informativo)</h2>
    <div class="info">
        El score de blur está normalizado por resolución para que sea comparable
        entre imágenes de 224px y de 4000px. <b>No excluye imágenes automáticamente.</b>
        Úsalo como referencia en la revisión manual de las grillas.
        Score bajo = imagen posiblemente borrosa.
    </div>"""

    # ── Sección grillas ──────────────────────────────────────────────────────
    cls_colors = {
        "late_blight":    "#e8f4f8",
        "early_blight":   "#e3f2fd",
        "leafroll_virus": "#e8f5e9",
        "mosaic_virus":   "#f1f8e9",
        "bacterial_wilt": "#fef9e7",
        "nematode":       "#fdf2f8",
        "pest":           "#fce4ec",
        "healthy":        "#e8f5e9",
        "__REVISAR__":    "#fff8e1",
        "__DESCONOCIDO__":"#fdecea",
    }

    grids_html = f"""
    <h2>Grillas de revisión visual ({N_MUESTRAS} muestras por clase por dataset)</h2>
    <div class="warn">
        <b>Qué revisar:</b><br>
        • Grillas <b>__REVISAR__</b>: decidir a qué clase pertenece cada imagen de fungi/virus.<br>
        • Grillas de DS1 y DS6: identificar imágenes de planta completa irrecuperables.<br>
        • Cualquier imagen que no corresponda a la clase indicada → anotar en el Excel.
    </div>"""

    ds_order = sorted(set(ds for ds, _ in grids.keys()))
    for ds in ds_order:
        grids_html += f"<h3 style='margin-top:28px'>{ds}</h3>"
        cls_order = sorted(set(cls for d, cls in grids.keys() if d == ds))

        # Primero clases objetivo, luego __REVISAR__, luego __DESCONOCIDO__
        priority = TARGET_CLASSES + ["__REVISAR__", "__DESCONOCIDO__"]
        cls_sorted = [c for c in priority if c in cls_order] + \
                     [c for c in cls_order if c not in priority]

        for cls in cls_sorted:
            imgs = grids.get((ds, cls), [])
            if not imgs:
                continue
            bg = cls_colors.get(cls, "#f9f9f9")
            revisar_badge = ""
            if cls == "__REVISAR__":
                revisar_badge = """<span style="background:#e65100;color:white;
                    font-size:11px;padding:2px 8px;border-radius:3px;margin-left:8px;">
                    CLASIFICACIÓN MANUAL REQUERIDA</span>"""

            grids_html += f"""
            <div style="margin-bottom:16px;padding:10px;background:{bg};
                        border-radius:6px;border:1px solid #ddd;">
                <p style="margin:0 0 6px;font-weight:bold;font-size:13px;">
                    {cls} {revisar_badge}
                    <span style="font-weight:normal;font-size:11px;color:#666;">
                        ({len(imgs)} de {min(N_MUESTRAS, len(imgs))} mostradas)
                    </span>
                </p>
                <div style="display:flex;flex-wrap:wrap;gap:3px;">"""

            for b64 in imgs:
                grids_html += f"""<img src="data:image/jpeg;base64,{b64}"
                    style="width:88px;height:88px;object-fit:cover;
                           border-radius:3px;border:1px solid #ccc;" title="{cls}">"""

            grids_html += "</div></div>"

    # ── Sección pasos siguientes ─────────────────────────────────────────────
    next_steps = """
    <h2>Checklist — antes de pasar al Paso 2</h2>
    <ol>
        <li>✔ Confirmar exclusión de datasets subconjunto (DS1 y DS7 si se confirman).</li>
        <li>Revisar grillas <b>__REVISAR__</b> y asignar cada imagen a su clase correcta
            o confirmar exclusión.</li>
        <li>Revisar grillas de DS1 y DS6: marcar imágenes de planta completa
            irrecuperables en el Excel.</li>
        <li>Revisar hoja <b>Error_etiquetado</b> del Excel: corregir imágenes
            asignadas a dos clases distintas.</li>
        <li>Revisar hoja <b>Intra_clase</b> del Excel: eliminar copias exactas
            dentro del mismo dataset.</li>
        <li>Con esas decisiones tomadas → ejecutar Paso 2 (consolidación).</li>
    </ol>"""

    # ── HTML final ────────────────────────────────────────────────────────────
    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Auditoría Completa — Papa CNN</title>
<style>
  body  {{ font-family: Arial, sans-serif; margin: 24px; font-size: 14px;
          line-height: 1.5; }}
  h1   {{ color: #2c5f2e; }}
  h2   {{ color: #333; border-bottom: 2px solid #2c5f2e; padding-bottom: 4px;
          margin-top: 32px; }}
  h3   {{ color: #555; margin-top: 18px; }}
  table {{ border-collapse: collapse; width: 100%; margin-bottom: 16px;
           font-size: 13px; }}
  th   {{ background: #2c5f2e; color: white; padding: 8px; text-align: left; }}
  td   {{ border: 1px solid #ddd; padding: 6px; vertical-align: top; }}
  .warn {{ background: #fff3cd; padding: 10px 14px;
           border-left: 4px solid #ffc107; border-radius: 3px;
           margin-bottom: 12px; }}
  .info {{ background: #d1ecf1; padding: 10px 14px;
           border-left: 4px solid #17a2b8; border-radius: 3px;
           margin-bottom: 12px; }}
</style>
</head>
<body>
<h1>🥔 Auditoría Completa de Datasets — Clasificación de Enfermedades de Papa</h1>
<p>
  <b>Generado:</b> {timestamp} &nbsp;|&nbsp;
  <b>Datasets:</b> {len(DATASETS)} &nbsp;|&nbsp;
  <b>Registros totales:</b> {len(records):,} &nbsp;|&nbsp;
  <b>Corruptas:</b> {len(corrupt_list)} &nbsp;|&nbsp;
  <b>Método hash:</b> {'pHash perceptual' if USE_PHASH else 'MD5'}
</p>

<h2>Resumen por clase objetivo</h2>
<div class="info">
  Criterio de viabilidad: ≥1.000 imágenes = OK, 400–999 = Aceptable, &lt;400 = Crítico.
  Total entrenable (8 clases): <b>{total_entrenables:,}</b> imágenes.
</div>
<table>
  <tr>
    <th>Clase</th><th>Imágenes</th>
    <th>% del total entrenable</th><th>Estado</th>
  </tr>
  {summary_rows}
</table>

{subset_section}
{dup_section}
{blur_section}
{grids_html}
{next_steps}
</body>
</html>"""

    REPORT_HTML.write_text(html, encoding="utf-8")
    log.info(f"[OK] Reporte HTML guardado: {REPORT_HTML}")


# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════

def main():
    log.info("=" * 60)
    log.info("PASO 2 — Auditoría Completa de Datasets")
    log.info("=" * 60)

    # 1. Escaneo
    log.info("\n[1] Escaneando datasets...")
    records, corrupt_list = scan_all_datasets()

    # 2. Duplicados cruzados
    log.info("\n[2] Detectando duplicados cruzados...")
    duplicates = detect_cross_duplicates(records)

    # 3. Subconjuntos
    log.info("\n[3] Detectando subconjuntos...")
    subsets = detect_subsets(records)

    # 4. Grillas
    log.info(f"\n[4] Generando grillas ({N_MUESTRAS} muestras por clase/dataset)...")
    grids = generate_grids(records)

    # 5. CSV
    log.info("\n[5] Guardando CSV de inventario...")
    save_csv(records)

    # 6. Excel duplicados
    log.info("\n[6] Guardando Excel de duplicados...")
    save_duplicates_excel(duplicates)

    # 7. Reporte HTML
    log.info("\n[7] Generando reporte HTML...")
    save_html_report(records, corrupt_list, duplicates, subsets, grids)

    # ── Resumen consola ──────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("RESUMEN FINAL")
    print("=" * 60)

    class_totals = defaultdict(int)
    for r in records:
        class_totals[r["mapped_class"]] += 1

    print("\nClases objetivo:")
    for cls in TARGET_CLASSES:
        n = class_totals.get(cls, 0)
        bar = "█" * min(n // 100, 40)
        st = "✔" if n >= 1000 else ("⚠" if n >= 400 else "✘")
        print(f"  {st} {cls:<22} {n:>5}  {bar}")

    print(f"\n  ⚠ __REVISAR__  (fungi/virus)  {class_totals.get('__REVISAR__', 0):>5}")
    print(f"  ✘ __DESCONOCIDO__             {class_totals.get('__DESCONOCIDO__', 0):>5}")

    if subsets:
        print("\nSubconjuntos — recomendación de exclusión:")
        for s in subsets:
            print(f"  → EXCLUIR {s['subset']} ({s['pct']}% en {s['superset']})")

    print(f"\nArchivos generados en: {OUTPUT_DIR.resolve()}")
    print(f"  {REPORT_HTML.name}")
    print(f"  {INVENTORY_CSV.name}")
    print(f"  {DUPLICATES_XLS.name}")
    print(f"  auditoria.log")


if __name__ == "__main__":
    main()
